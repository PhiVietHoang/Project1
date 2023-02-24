import time
from array import array
from asyncio.windows_events import NULL
from cmath import log
from operator import index
from pathlib import Path
from pickle import NONE
import tkinter as tk
import customtkinter as customTkinter
import tkinter.scrolledtext as st
from tkinter import ANCHOR, BROWSE, CENTER, YES, INSERT, END, ttk
from tkinter.messagebox import NO
from turtle import onclick, width

import pandas
import pandas as pd
import numpy as np
from sklearn.metrics import roc_curve, auc, RocCurveDisplay
import openpyxl
import classifier as clf
from tkinter import W, IntVar, StringVar, filedialog as fd
import matplotlib.pyplot as plt
import tach_file
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg)
# Implement the default Matplotlib key bindings.
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import os

np.seterr(divide='ignore', invalid='ignore')
global L_pred, L_test

# Configure windows
window = customTkinter.CTk()

customTkinter.set_appearance_mode("light")  # Modes: system (default), light, dark
customTkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

window.title('Project 1')
window.geometry('1000x600+300+200')
# window.configure(bg='#fff')
# window.resizable(False, False)


headerFrame = tk.Frame(window, borderwidth=1, relief='solid', padx=16, pady=4)
headerFrame.place(x=0, y=0, relwidth=1, relheight=0.07)

# Configure Input Frame
inputFrame = tk.Frame(window, borderwidth=1, relief='solid', padx=16, pady=16)
inputFrame.place(x=0, rely=0.07, relheight=0.86, relwidth=0.65)

propFrame = tk.Frame(master=inputFrame)

# Configure Output Frame
outputFrame = tk.Frame(window, borderwidth=1, relief='solid', padx=16, pady=16)
outputFrame.place(relx=0.65, rely=0.07, relheight=(1 - 0.14), relwidth=0.35)

footerFrame = tk.Frame(window, borderwidth=1, relief='solid', padx=16, pady=4)
footerFrame.place(x=0, rely=(1 - 0.07), relwidth=1, relheight=0.07)

excelFile = pd.DataFrame
filename = ''

fileLabel = customTkinter.CTkLabel(headerFrame, text="")
fileLabel.place(relx=0.25, rely=0.04)

buttonChooseFile = customTkinter.CTkButton(headerFrame, text="Choose File!",
                                           command=lambda: openFile())
buttonChooseFile.place(relx=0.05, rely=0.03)

accuracyLabel = customTkinter.CTkLabel(outputFrame, text='Accuracy: ', anchor=W)
# accuracyLabel.place(rely=0.01)
accuracyLabel.grid(row=0, column=0, sticky='ew')

weightLabel = customTkinter.CTkLabel(outputFrame, text='Weight: ', anchor=W)
# weightLabel.place(rely=0.06)
weightLabel.grid(row=2, column=0, sticky='ew')

timeLabel = customTkinter.CTkLabel(outputFrame, text='Time: ', anchor=W)
timeLabel.grid(row=1, column=0, sticky='ew')

weightResult = st.ScrolledText(outputFrame, width=10, height=5)
# weightResult.place(rely=0.11)
weightResult.grid(row=3, column=0, sticky='ew')

outputFrame.columnconfigure(0, weight=1)
outputFrame.columnconfigure(1, weight=1)
outputFrame.rowconfigure(4, weight=1)


def prob_label(R, R_Lp, classes, w, ruler):
    prob = np.zeros((R.shape[0], classes.shape[0]))
    y, n = clf.fuzzify(R, R)
    y_Lp, n_Lp = clf.fuzzify(R_Lp, R)

    for i in range(R.shape[0]):
        dists = np.zeros(classes.shape[0])
        for j in range(classes.shape[0]):
            dists[j] = clf.distance_calc(y[i], n[i], y_Lp[j], n_Lp[j], w, ruler)
        prob[i] = np.exp(-dists) / np.exp(-dists).sum(axis=0)

    return prob


def reset_gui():
    global result_data, w
    detailBtn.configure(state="disabled")
    matrixBtn.configure(state="disabled")
    weightResult.configure(state='normal')
    weightResult.delete('1.0', END)
    result_data = pandas.DataFrame()
    w = pandas.DataFrame()


footerFrame.columnconfigure('all', weight=1)

splitbtn = customTkinter.CTkButton(footerFrame, text="Split", command=lambda: split_data())
splitbtn.grid(row=0, column=0, sticky="ew")

matrixBtn = customTkinter.CTkButton(footerFrame, text="Confusion Matrix", command=lambda: get_confusion_matrix()
                                    , state='disabled')
matrixBtn.grid(row=0, column=2, padx=5, sticky="ew")

detailBtn = customTkinter.CTkButton(footerFrame, text="Detailed Stats", command=lambda: get_detail_stats()
                                    , state="disabled")
detailBtn.grid(row=0, column=3, padx=5, sticky="ew")

exportBTN = customTkinter.CTkButton(footerFrame, text="Export", command=lambda: export_result())
exportBTN.grid(row=0, column=5, padx=5, sticky="e")

buttonProcess = customTkinter.CTkButton(footerFrame, text="Process", command=lambda: processingData())
# buttonProcess.place(relwidth = ,relx=0.45, rely=0,)
buttonProcess.grid(row=0, column=1, sticky='ew', padx=5)


def openFile():
    global filename
    filename = fd.askopenfilename(filetypes=[("Data files",
                                              ["*.xlsx", "*.csv", "*xls"])]).replace('\\', '/')
    global excelFile

    if filename[-5:] == ".xlsx" or filename[-4:] == ".xls":
        fileLabel.configure(text="File: " + filename.split('/')[-1])
        excelFile = pd.read_excel(filename)
        loadData()
    elif filename[-4:] == ".csv":
        fileLabel.configure(text="File: " + filename.split('/')[-1])
        excelFile = pd.read_csv(filename)
        loadData()
    else:
        fileLabel.configure(text="Not a correct data file")


do_label = ['Euler', 'Hamming (2 hàm)', 'Hamming (3 hàm)', 'Mahanta', 'Ngân']
do_values = ['eu', 'ha2', 'ha3', 'ma', 'ng']

vars = []
doDoVal = IntVar()

top_level_in_use = []
result_data = pandas.DataFrame()
w = pandas.DataFrame()


def loadData():
    for item in propFrame.winfo_children():
        item.destroy()
    global columns, l_values
    columns = list(excelFile.columns.values)
    l_values = (
        list(dict.fromkeys(list(excelFile[excelFile.columns.values[-1]].values))))

    loadTable(excelFile, columns)

    def selectColumns(index, value):
        if (columns[index] == value):
            columns[index] = None
        else:
            columns[index] = value

        loadTable(excelFile, [item for item in columns if item != None])

    def selectLValues(index, value):
        if (l_values[index] == value):
            l_values[index] = None
        else:
            l_values[index] = value

    columnsFrame = tk.Frame(inputFrame)
    columnsFrame.place(relwidth=1, rely=0.76)


    for index, prop in enumerate(columns[:-1]):
        vars.append(StringVar())
        vars[-1].set(1)
        radioButton = customTkinter.CTkCheckBox(columnsFrame, text=columns[index], variable=vars[-1], checkbox_width=18,
                                                checkbox_height=18,
                                                command=lambda value=columns[index], index=index: selectColumns(index,
                                                                                                                value))
        radioButton.pack(side='left')

    doDoFrame = tk.Frame(inputFrame)
    doDoFrame.place(relwidth=1, rely=0.82)
    for index, doValue in enumerate(do_label):
        radioButton = customTkinter.CTkRadioButton(
            doDoFrame, text=doValue, variable=doDoVal, value=(index + 1), radiobutton_height=18, radiobutton_width=18,
            width=100)
        radioButton.pack(side='left')
    reset_gui()


def loadTable(excelFile, columns):
    tree = ttk.Treeview(inputFrame, columns=columns, show='headings')

    isTrain = columns.count('Prediction') > 0
    for index, item in enumerate(columns):
        tree.heading(item, text=item, anchor=CENTER)
        tree.column('#' + str(index), width=10, stretch=YES)
    tree.column('#' + str(len(columns)), width=0)

    showData = excelFile[columns].values

    for data in showData:
        if (isTrain and data[-1] != data[-2]):
            tree.insert('', tk.END, values=list(data), tags="wrongPrediction")
        else:
            tree.insert('', tk.END, values=list(data))

    if (isTrain):
        tree.tag_configure('wrongPrediction', background='#f2ec6d')

    tree.grid(row=0, column=0, sticky='we')
    tree.place(y=0, relwidth=1.0, relheight=0.56, relx=0)


def processingData():
    if not excelFile.empty:
        selectedColumns = [item for item in columns[:-1] if item != None]

        train_data = excelFile[selectedColumns].values
        w0 = 1 / len(selectedColumns)
        st = time.time()
        weight_train = clf.weight_train(w0, train_data)

        L_train = excelFile.to_numpy()[:, -1]
        classes = np.unique(excelFile.to_numpy()[:, -1])
        center = clf.center_find(train_data, L_train, classes)
        doDo = do_values[doDoVal.get() - 1]
        print(doDo)
        global L_pred, result_data, w

        w = clf.early_stopping(weight_train, L_train, train_data, center, classes, doDo)
        L_pred = clf.labeling(train_data, center, classes, w, doDo)
        accuracy = clf.accuracy(L_pred, L_train)
        et = time.time()
        timeEx = round((et - st) * 1000, 2)

        test = excelFile.to_numpy()
        L_test = test[:, len(excelFile.columns) - 1].astype('<U8')

        ACC = clf.acc(L_pred, L_test, classes)
        SEN = clf.sen(L_pred, L_test, classes)
        SPEC = clf.spec(L_pred, L_test, classes)
        PRE = clf.pre(L_pred, L_test, classes)
        F1 = clf.f1_score(L_pred, L_test, classes)

        result_data = pandas.DataFrame([ACC, SEN, SPEC, PRE, F1]).transpose()
        result_data.index = classes
        result_data.columns = ["Accuracy", "Sensitivity", "Specificity", "Precision", "F1 score"]
        result_data = result_data.round(2)

        x = [item for item in columns if item != None]
        x.append(('Prediction'))
        loadTable(pd.concat([excelFile, pd.Series(
            L_pred, name='Prediction')], axis=1), list(x))

        accuracyLabel.configure(text="Accuracy: " + str(round(accuracy, 5)))


        w = round(pd.Series(weight_train[-1], index=selectedColumns), 5)
        weightStr = w.to_string(index=True)
        weightResult.insert(INSERT, weightStr)
        weightResult.configure(state='disabled')
        timeLabel.configure(text="Time: " + str(timeEx) + " ms")

        x = np.arange(weight_train.shape[0])
        y = np.zeros_like(x, dtype='float')
        for i in range(x.shape[0]):
            y[i] = clf.accuracy(clf.labeling(train_data, center, classes, weight_train[i], doDo),
                                L_train)  # Day la doan khien chuong trinh chay lau
        fig = plt.figure(figsize=(3.6, 3.4), dpi=100)
        fig.add_subplot(111).plot(x, y)
        plt.xlabel("Iteration")
        plt.ylabel("Accuracy")
        plt.title("Accuracy graph")
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=outputFrame)  # A tk.DrawingArea.
        canvas.draw()
        canvas.get_tk_widget().grid(row=4, sticky="news", column=0, columnspan=2, pady=5, padx=5)
        detailBtn.configure(state="active")
        matrixBtn.configure(state="active")



# print(excelFile)
def split_data():
    if not excelFile.empty:
        dfData = excelFile
        listColumns = dfData.columns.to_list()
        train = pd.DataFrame()
        validate = pd.DataFrame()
        test = pd.DataFrame()
        columnsLabelName = listColumns[len(listColumns) - 1]
        listLabel = dfData[columnsLabelName].unique().tolist()

        for y in listLabel:
            dfy = dfData[dfData[columnsLabelName] == y]
            train1, validate1, test1 = tach_file.train_validate_test_split(dfy)
            train = pd.concat([train, train1])
            validate = pd.concat([validate, validate1])
            test = pd.concat([test, test1])

        path = Path(filename)
        train.to_csv(f'Train_' + path.name, index=False)
        validate.to_csv(f'Validate_' + path.name, index=False)
        test.to_csv(f'Test_' + path.name, index=False)


def create_toplevel(rowname, colname, def_matrix):
    if len(top_level_in_use) == 1:
        top_level_in_use[0].destroy()
        top_level_in_use.pop()

    top = customTkinter.CTkToplevel(window)
    top_level_in_use.append(top)
    top.title(" Defuse Matrix")

    top_label = customTkinter.CTkLabel(top, text="Classes", height=3, width=10)
    top_label.grid(row=0, column=0)

    row = len(rowname)
    col = len(colname)
    for i in range(col):
        top_label1 = customTkinter.CTkLabel(top, text=colname[i], height=30, width=80)
        top_label1.grid(row=0, column=i + 1)
    for j in range(row):
        top_label2 = customTkinter.CTkLabel(top, text=rowname[j], height=30, width=80)
        top_label2.grid(row=j + 1, column=0)

    for i in range(row):
        for j in range(col):
            top_label1 = customTkinter.CTkLabel(top, text=def_matrix[i][j], height=30, width=80)
            top_label1.grid(row=i + 1, column=j + 1)

    top.mainloop()


def get_confusion_matrix():
    global L_pred
    if not excelFile.empty:
        test = excelFile.to_numpy()
        L_test = test[:, len(excelFile.columns) - 1].astype('<U8')
        classes = np.unique(L_pred)
        TP, FP, TN, FN = clf.value_of_confusion_matrix(L_pred, L_test, classes)
        matrix = np.array([TP, FP, TN, FN]).T.reshape(len(classes), 4)
        create_toplevel(classes, ["TP", "FP", "TN", "FN"], matrix)


def get_detail_stats():
    if not result_data.empty:
        classes = np.unique(excelFile.to_numpy()[:, -1])
        col_names = ["ACC", "SEN", "SPEC", "PRE", "F1"]
        create_toplevel(classes, col_names, result_data.values.tolist())


def export_result():
    if not result_data.empty:
        file_location = os.path.splitext(filename)[0] + "_Result.xlsx"
        if not os.path.exists(file_location):
            with pandas.ExcelWriter(file_location) as writer:
                result_data.to_excel(writer, sheet_name="Result")
                cell = writer.sheets['Result'].cell(row=1, column=7)
                cell.value = do_label[doDoVal.get()-1]
                w.to_excel(writer, sheet_name="Weight")
        else:
            with pandas.ExcelWriter(file_location,mode='a',if_sheet_exists='overlay') as writer:
                startrow1 = writer.sheets['Result'].max_row
                cell = writer.sheets['Result'].cell(row=startrow1+1, column=7)
                cell.value = do_label[doDoVal.get()-1]
                result_data.to_excel(writer, sheet_name="Result", startrow=startrow1)
        plt.savefig(os.path.splitext(filename)[0] + "_graph.png")
        print("Exported to " + os.path.abspath(file_location))

window.mainloop()
